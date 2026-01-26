from openpyxl import load_workbook
import os
import json
from typing import Optional, Tuple, Dict, Any
from backend.database import get_connection


SHEET_PRECIOS = "NUEVA LISTA DE PRECIOS"

JSON_IN = "productos_precios.json"
JSON_OUT = "productos_precios.json"

TIPO_CAMBIO = float(os.getenv("TIPO_CAMBIO", "6.96"))

# Intentaremos usar el Excel subido por el panel
EXCEL_CANDIDATES = [
    "proveedor.xlsm",
    "proveedor.xlsx",
    "proveedor.xls",
    # fallback viejo:
    "NOTA DE PEDIDO VER. 8-12-2025 OFICIAL.xlsm",
]

ALLOWED_EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}

def to_float(x):
    try:
        if x is None:
            return None
        return float(x)
    except:
        return None

def _parse_discount_cell(v) -> Optional[float]:
    """
    Convierte valores tipo:
    - 0.2  -> 0.2
    - 20   -> 0.2
    - "20%" -> 0.2
    """
    if v is None:
        return None

    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        if s.endswith("%"):
            s = s[:-1].strip()
        try:
            v = float(s)
        except:
            return None

    try:
        v = float(v)
    except:
        return None

    if v > 1:
        v = v / 100.0

    if v < 0 or v > 0.95:
        return None

    return v

def _find_excel_path(base_dir: str) -> Tuple[Optional[str], list]:
    """
    Prioridad:
    1) Si app.py seteó EXCEL_FILE
    2) Candidatos conocidos
    """
    checked = []

    env_name = os.environ.get("EXCEL_FILE")
    if env_name:
        p = os.path.join(base_dir, env_name)
        checked.append(p)
        if os.path.exists(p):
            return p, checked

    for name in EXCEL_CANDIDATES:
        p = os.path.join(base_dir, name)
        checked.append(p)
        if os.path.exists(p):
            return p, checked

    return None, checked

def _calc_margen(costo_bs: float) -> float:
    # Tu regla de tramos (mantengo tu idea)
    if costo_bs < 30:
        return 0.45
    elif costo_bs < 80:
        return 0.35
    elif costo_bs < 200:
        return 0.28
    else:
        return 0.20

def actualizar_precios(descuento_proveedor: Optional[float] = None):
    """
    - Lee el Excel del proveedor
    - Lee descuento desde G6 (si descuento_proveedor es None)
    - Actualiza productos_precios.json
    - CREA productos nuevos si aparecen en Excel y no existían en JSON
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))

    excel_path, checked = _find_excel_path(base_dir)
    json_in_path = os.path.join(base_dir, JSON_IN)
    json_out_path = os.path.join(base_dir, JSON_OUT)

    if not excel_path:
        return {"ok": False, "error": "No encuentro el Excel del proveedor", "checked": checked}

    #if not os.path.exists(json_in_path):
        #return {"ok": False, "error": "No encuentro productos_precios.json", "path": json_in_path}

    ext = os.path.splitext(excel_path)[1].lower()
    if ext not in ALLOWED_EXCEL_EXTS:
        return {"ok": False, "error": f"Extensión Excel no soportada: {ext}"}

    # Abrimos en modo lectura para que sea MUCHO más rápido y no cargue macros
    wb = load_workbook(excel_path, data_only=True, read_only=True, keep_vba=False)

    # 1) Descuento desde G6 en HOJA PEDIDO
    ws_header = wb["HOJA PEDIDO"] if "HOJA PEDIDO" in wb.sheetnames else wb.active
    descuento_excel = _parse_discount_cell(ws_header["G6"].value)

    if descuento_proveedor is None:
        descuento_proveedor = descuento_excel if descuento_excel is not None else 0.20
    else:
        descuento_proveedor = _parse_discount_cell(descuento_proveedor) or float(descuento_proveedor)

    # 2) Hoja de precios
    if SHEET_PRECIOS not in wb.sheetnames:
        return {"ok": False, "error": f"No existe la hoja '{SHEET_PRECIOS}'", "sheets": wb.sheetnames}

    ws = wb[SHEET_PRECIOS]

    # 3) Leer Excel completo (desde fila 3)
    excel_by_code: Dict[str, Dict[str, Any]] = {}
    filas_excel = 0

    for r in ws.iter_rows(min_row=3, values_only=True):
        # ESTRUCTURA REAL de tu "NUEVA LISTA DE PRECIOS" (según tu Excel):
        # A: productCode (ej TR-15725)
        # B: CODIGO (ej 15725)  <-- ESTE ES EL CODE QUE COINCIDE CON tu JSON
        # C: CO (ej TR-)
        # F: ubicación
        # G: DESCRIPCIÓN
        # H: P/U (USD)
        # I: EMPAQUE (texto)
        # J: MARCA
        # L: almacen (CENTRAL, etc.)

        product_code = r[0]
        codigo = r[1]
        co = r[2]
        ubicacion = r[5]
        descripcion = r[6]
        usd_unit = r[7]
        empaque = r[8]
        marca = r[9]
        almacen = r[11]

        if codigo is None:
            continue

        code_raw = str(codigo).strip()
        if code_raw.endswith(".0"):
            code_raw = code_raw[:-2]
        code = code_raw

        if not code:
            continue

        if not code.isdigit():
            continue

        usd_u = to_float(usd_unit)

        # si no hay precio USD, no sirve para actualizar
        if usd_u is None:
            continue

        filas_excel += 1

        excel_by_code[code] = {
            "code": code,
            "productCode": str(product_code).strip() if product_code else None,
            "co": str(co).strip() if co else None,
            "location": str(ubicacion).strip() if ubicacion is not None else None,
            "description": str(descripcion).strip() if descripcion else "",
            "package": str(empaque).strip() if empaque else "",
            "brand": str(marca).strip().lower() if marca else "",
            "warehouse": str(almacen).strip() if almacen else "",
            "usd_price_unit": usd_u,
            # en este Excel NO viene docena ni Bs directo
            "usd_price_docena": None,
            "bs_price_proveedor": None,
        }



        # =========================
    # BD: cargar catálogo actual
    # =========================
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT code, data FROM productos_catalogo")
    rows = cur.fetchall() or []

    db_by_code = {}
    for r in rows:
        c = str(r.get("code") or "").strip()
        if c:
            db_by_code[c] = r.get("data") or {}

    db_codes = set(db_by_code.keys())
    excel_codes = set(excel_by_code.keys())

    # En BD pero no en Excel (missing)
    en_db_no_en_excel = sorted(list(db_codes - excel_codes))

    actualizados = 0
    nuevos_detectados = []

    now = datetime.utcnow().isoformat()

    # =========================
    # Actualizar existentes
    # =========================
    for code, ex in excel_by_code.items():
        if code not in db_by_code:
            # nuevo detectado (NO insertar automático)
            nuevos_detectados.append({
                "code": code,
                "description": ex.get("description", ""),
                "brand": ex.get("brand", ""),
                "usd_price_unit": ex.get("usd_price_unit"),
            })
            continue

        p = dict(db_by_code[code])  # base actual (preserva fields existentes)

        # Actualizar datos base desde Excel
        p["code"] = code
        p["productCode"] = ex.get("productCode")
        p["co"] = ex.get("co")
        p["location"] = ex.get("location")
        p["warehouse"] = ex.get("warehouse")
        p["description"] = ex.get("description", p.get("description", ""))
        p["brand"] = (ex.get("brand") or p.get("brand") or "").strip().lower()

        usd_unit = ex.get("usd_price_unit")
        p["usd_price_unit"] = usd_unit

        # Mantener si ya existía
        usd_docena = p.get("usd_price_docena")
        p["usd_price_docena"] = usd_docena

        # Guardar descuento proveedor aplicado (G6)
        p["proveedor_descuento"] = float(descuento_proveedor)

        # --- aquí debes usar tu misma lógica existente de cálculo de precios ---
        # IMPORTANTE: reutiliza las funciones/márgenes que ya tienes en tu script.
        # Normalmente tu script termina calculando:
        # p["usd_price_web"], p["bs_price_descuento25"], p["bs_price_web"], sale_label, box_qty, etc.
        #
        # Si tu script ya tiene una función para calcular esto, LLÁMALA aquí.
        # Si no, dime y lo integramos sin inventar.

        # Persistir en BD (upsert)
        cur.execute(
            """
            INSERT INTO productos_catalogo (code, data, updated_at)
            VALUES (%s, %s::jsonb, %s)
            ON CONFLICT (code)
            DO UPDATE SET data = EXCLUDED.data, updated_at = EXCLUDED.updated_at
            """,
            (code, json.dumps(p, ensure_ascii=False), now)
        )
        actualizados += 1

    conn.commit()
    conn.close()

    return {
        "ok": True,
        "actualizados": actualizados,
        "filas_excel_validas": filas_excel,
        "descuento_proveedor": float(descuento_proveedor),
        "en_json_no_en_excel": en_db_no_en_excel,   # compat: el app.py lo mapea a missing
        "nuevos_detectados": nuevos_detectados
    }
